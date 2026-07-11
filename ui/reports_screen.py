"""
reports_screen.py — Reportes PAE Control

- KPI cards: registros / strikes / estudiantes afectados
- Tabs: Top 25 semana / Top 25 mes / Asistencia / Lista de espera / Log altas-bajas
- Exportar CSV (tabs de strikes) o Excel (asistencia)
- Auto-refresh al entrar a la pantalla
"""

import csv
import os
import subprocess
from datetime import datetime, date, timedelta

from PyQt6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel,
    QTableWidget, QTableWidgetItem, QHeaderView,
    QTabWidget, QAbstractItemView, QFrame, QDateEdit,
    QPushButton, QSizePolicy
)
from PyQt6.QtCore import Qt, QDate
from PyQt6.QtGui import QColor, QFont

import db
import utils
from ui.theme   import C, sound
from ui.widgets import AButton, SectionHeader, StatCard, ToastBanner, ConfirmPanel, make_table_copyable


def _rank_color(rank: int) -> str:
    if rank <= 5:   return C.RED
    if rank <= 10:  return C.AMBER
    if rank <= 15:  return C.GOLD_500
    return C.TEXT3


def _export_excel(rows: list, headers: list, sheet_name: str, filepath: str):
    """Genera un .xlsx con formato básico usando openpyxl."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        return False

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name

    HDR_FILL = PatternFill("solid", fgColor="1E3A5F")
    HDR_FONT = Font(bold=True, color="FFFFFF", size=11)
    ALT_FILL = PatternFill("solid", fgColor="1A2233")

    # Cabeceras
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = HDR_FONT
        cell.fill      = HDR_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Datos
    for r_idx, row in enumerate(rows, 2):
        fill = ALT_FILL if r_idx % 2 == 0 else None
        for c_idx, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=c_idx, value=val)
            if fill:
                cell.fill = fill
            cell.alignment = Alignment(vertical="center")

    # Ancho automático
    for col in ws.columns:
        max_w = max((len(str(c.value or "")) for c in col), default=8)
        ws.column_dimensions[col[0].column_letter].width = min(max_w + 4, 40)

    wb.save(filepath)
    return True


# ══════════════════════════════════════════════════════
#  TAB DE ASISTENCIA
# ══════════════════════════════════════════════════════

class _AsistenciaTab(QWidget):
    """
    Tab con reportes de asistencia diario / semanal / mensual.
    Incluye top de estudiantes más responsables y exportación a Excel.
    """

    _PERIODOS = [
        ("Hoy",         "hoy"),
        ("Esta semana", "semana"),
        ("Este mes",    "mes"),
        ("Personalizado","custom"),
    ]

    def __init__(self, parent=None):
        super().__init__(parent)
        self._desde: date = date.today()
        self._hasta: date = date.today()
        self._build_ui()
        self._set_periodo("semana")

    def _build_ui(self):
        self.setStyleSheet(f"background: {C.BG};")
        root = QVBoxLayout(self)
        root.setContentsMargins(20, 16, 20, 16)
        root.setSpacing(12)

        # ── Selector de período ──────────────────────
        period_row = QHBoxLayout()
        period_row.setSpacing(6)
        self._period_btns: dict[str, QPushButton] = {}
        for label, key in self._PERIODOS:
            btn = QPushButton(label)
            btn.setFixedHeight(30)
            btn.setCursor(Qt.CursorShape.PointingHandCursor)
            btn.setProperty("period_key", key)
            btn.clicked.connect(lambda _, k=key: self._set_periodo(k))
            period_row.addWidget(btn)
            self._period_btns[key] = btn
        period_row.addSpacing(16)

        # Date pickers (solo visibles en "custom")
        self._lbl_desde = QLabel("Desde")
        self._lbl_desde.setStyleSheet(f"font-size: 11px; color: {C.TEXT3}; background: transparent;")
        self._date_desde = QDateEdit()
        self._date_desde.setCalendarPopup(True)
        self._date_desde.setDisplayFormat("dd/MM/yyyy")
        self._date_desde.setDate(QDate.currentDate())
        self._date_desde.setFixedHeight(30)

        self._lbl_hasta = QLabel("Hasta")
        self._lbl_hasta.setStyleSheet(f"font-size: 11px; color: {C.TEXT3}; background: transparent;")
        self._date_hasta = QDateEdit()
        self._date_hasta.setCalendarPopup(True)
        self._date_hasta.setDisplayFormat("dd/MM/yyyy")
        self._date_hasta.setDate(QDate.currentDate())
        self._date_hasta.setFixedHeight(30)

        for w in (self._lbl_desde, self._date_desde, self._lbl_hasta, self._date_hasta):
            period_row.addWidget(w)
            w.setVisible(False)
        self._custom_widgets = [self._lbl_desde, self._date_desde,
                                self._lbl_hasta, self._date_hasta]

        btn_apply = QPushButton("Aplicar")
        btn_apply.setFixedHeight(30)
        btn_apply.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_apply.clicked.connect(self._apply_custom)
        btn_apply.setVisible(False)
        period_row.addWidget(btn_apply)
        self._btn_apply = btn_apply

        period_row.addStretch()

        # Botón Excel
        btn_excel = QPushButton("↓  Exportar Excel")
        btn_excel.setFixedHeight(30)
        btn_excel.setCursor(Qt.CursorShape.PointingHandCursor)
        btn_excel.clicked.connect(self._export_excel)
        btn_excel.setStyleSheet(f"""
            QPushButton {{
                background: {C.GREEN_DIM}; color: {C.GREEN};
                border: 1.5px solid {C.GREEN}55; border-radius: 8px;
                padding: 0 14px; font-size: 12px; font-weight: 600;
            }}
            QPushButton:hover {{ background: {C.GREEN}22; }}
        """)
        period_row.addWidget(btn_excel)

        # Estilo de botones de período
        _inactive = f"""
            QPushButton {{
                background: {C.SURFACE2}; color: {C.TEXT2};
                border: 1px solid {C.BORDER}; border-radius: 8px;
                padding: 0 14px; font-size: 12px; font-weight: 500;
            }}
            QPushButton:hover {{ background: {C.NAVY_700}; color: {C.TEXT}; }}
        """
        for btn in self._period_btns.values():
            btn.setStyleSheet(_inactive)
        self._inactive_style = _inactive

        root.addLayout(period_row)

        # ── Barra de acento — conector visual entre selector y contenido ──
        self._accent_bar = QFrame()
        self._accent_bar.setFixedHeight(3)
        self._accent_bar.setStyleSheet(f"background: {C.BLUE}; border-radius: 2px;")
        root.addWidget(self._accent_bar)

        # ── KPI strip ───────────────────────────────
        kpi_row = QHBoxLayout()
        kpi_row.setSpacing(10)
        self._kpi_registros = StatCard("Total registros",     "—", accent=C.NAVY_400)
        self._kpi_unicos    = StatCard("Estudiantes únicos",  "—", accent=C.BLUE)
        self._kpi_dias      = StatCard("Días con servicio",   "—", accent=C.GREEN)
        for card in (self._kpi_registros, self._kpi_unicos, self._kpi_dias):
            card.setMinimumHeight(78)
            kpi_row.addWidget(card)
        root.addLayout(kpi_row)

        # ── Tabla ────────────────────────────────────
        self._lbl_period = QLabel("")
        self._lbl_period.setStyleSheet(
            f"font-size: 12px; font-weight: 600; color: {C.TEXT2}; background: transparent;"
        )
        root.addWidget(self._lbl_period)

        self._tabla = QTableWidget()
        self._tabla.setColumnCount(6)
        self._tabla.setHorizontalHeaderLabels(
            ["#", "RUN", "Apellidos, Nombres", "Curso", "Comidas", "Días asistidos"]
        )
        self._tabla.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        self._tabla.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._tabla.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        self._tabla.verticalHeader().setVisible(False)
        self._tabla.setAlternatingRowColors(True)
        self._tabla.setShowGrid(False)
        self._tabla.verticalHeader().setDefaultSectionSize(36)
        self._tabla.setStyleSheet(f"""
            QTableWidget {{
                background: {C.SURFACE};
                border: 1.5px solid {C.BORDER};
                border-left: 3px solid {C.BLUE};
                border-radius: 10px; outline: none;
            }}
            QTableWidget::item {{
                padding: 6px 12px; border: none; font-size: 12px; color: {C.TEXT2};
            }}
            QTableWidget::item:selected {{ background: {C.NAVY_700}; color: {C.TEXT}; }}
            QTableWidget::item:alternate {{ background: {C.SURFACE2}; }}
        """)
        root.addWidget(self._tabla, stretch=1)

    # ── Períodos ──────────────────────────────────────

    def _set_periodo(self, key: str):
        hoy = date.today()
        custom = key == "custom"

        for w in self._custom_widgets:
            w.setVisible(custom)
        self._btn_apply.setVisible(custom)

        _active = f"""
            QPushButton {{
                background: {C.BLUE_DIM}; color: {C.BLUE};
                border: 1.5px solid {C.BLUE}55;
                border-bottom: 3px solid {C.BLUE};
                border-radius: 8px 8px 0 0;
                padding: 0 14px; font-size: 12px; font-weight: 700;
            }}
        """
        for k, btn in self._period_btns.items():
            btn.setStyleSheet(_active if k == key else self._inactive_style)

        # Actualizar barra de acento
        self._accent_bar.setStyleSheet(f"background: {C.BLUE}; border-radius: 2px;")

        if key == "hoy":
            self._desde = self._hasta = hoy
        elif key == "semana":
            self._desde = hoy - timedelta(days=hoy.weekday())
            self._hasta = hoy
        elif key == "mes":
            self._desde = hoy.replace(day=1)
            self._hasta = hoy
        elif key == "custom":
            return   # espera que el usuario presione Aplicar

        self._load()

    def _apply_custom(self):
        d = self._date_desde.date()
        h = self._date_hasta.date()
        self._desde = date(d.year(), d.month(), d.day())
        self._hasta = date(h.year(), h.month(), h.day())
        self._load()

    # ── Datos ─────────────────────────────────────────

    def _load(self):
        desde_s = self._desde.isoformat()
        hasta_s = self._hasta.isoformat()

        self._lbl_period.setText(
            f"Período:  {utils.format_fecha_display(desde_s)}  →  {utils.format_fecha_display(hasta_s)}"
        )

        # KPIs
        try:
            rows = db.get_asistencia_periodo(desde_s, hasta_s)
            dias = db.get_dias_con_registros(desde_s, hasta_s)
            total_reg = sum(r["total_comidas"] for r in rows)
            unicos    = len(rows)
        except Exception:
            rows, dias, total_reg, unicos = [], [], 0, 0

        self._kpi_registros.set_value(str(total_reg))
        self._kpi_registros.flash()
        self._kpi_unicos.set_value(str(unicos))
        self._kpi_unicos.flash(C.BLUE)
        self._kpi_dias.set_value(str(len(dias)))
        self._kpi_dias.flash(C.GREEN)

        # Tabla
        self._tabla.setRowCount(len(rows))
        for i, row in enumerate(rows):
            run_fmt   = utils.run_display(row["run"])
            apellidos = f"{row['apellido_paterno'] or ''} {row['apellido_materno'] or ''}".strip()
            nombres   = row["nombres"] or ""
            nombre_full = f"{apellidos}, {nombres}".strip(", ")
            curso     = row["curso"] or ""
            comidas   = row["total_comidas"]
            dias_asi  = row["dias_distintos"]

            # Color según participación (verde = alta, ámbar = media, rojo = baja)
            pct = dias_asi / max(len(dias), 1)
            if pct >= 0.8:    color = C.GREEN
            elif pct >= 0.5:  color = C.AMBER
            else:             color = C.RED

            for col, val in enumerate([str(i+1), run_fmt, nombre_full, curso,
                                       str(comidas), str(dias_asi)]):
                item = QTableWidgetItem(val)
                item.setTextAlignment(Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft)
                if col in (4, 5):
                    item.setForeground(QColor(color))
                    f = item.font(); f.setBold(True); item.setFont(f)
                self._tabla.setItem(i, col, item)

        self._tabla.resizeColumnToContents(0)
        self._tabla.resizeColumnToContents(1)

    # ── Exportar Excel ────────────────────────────────

    def _export_excel(self):
        """
        Exporta un Excel con el registro dividido por día:
        - Cada fecha tiene su propia sección encabezada con la fecha
        - Dentro de cada sección: lista de estudiantes que asistieron ese día
          con todas las comidas que tuvieron
        - Al final: resumen de totales por día
        """
        desde_s = self._desde.isoformat()
        hasta_s = self._hasta.isoformat()

        export_dir = os.path.join(os.path.expanduser("~"), "pae_control", "exports")
        os.makedirs(export_dir, exist_ok=True)
        ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = os.path.join(export_dir, f"registro_{desde_s}_{hasta_s}_{ts}.xlsx")

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        except ImportError:
            # Fallback CSV agrupado por día
            self._export_csv_por_dia(desde_s, hasta_s, path.replace(".xlsx", ".csv"))
            return

        wb = openpyxl.Workbook()

        # ── Estilos ──────────────────────────────────────
        FILL_DATE   = PatternFill("solid", fgColor="1E3A5F")   # azul oscuro — fecha
        FILL_HDR    = PatternFill("solid", fgColor="2C4A6E")   # azul medio — columnas
        FILL_ALT    = PatternFill("solid", fgColor="F8F9FF")   # gris muy claro — fila par
        FILL_SUM    = PatternFill("solid", fgColor="E8F0FE")   # azul claro — resumen

        FONT_DATE   = Font(bold=True, color="FFFFFF", size=12)
        FONT_HDR    = Font(bold=True, color="FFFFFF", size=10)
        FONT_NORMAL = Font(size=10)
        FONT_BOLD   = Font(bold=True, size=10)
        FONT_SUM_H  = Font(bold=True, color="1E3A5F", size=10)

        thin = Side(style="thin", color="D1D1D6")
        border_bottom = Border(bottom=thin)

        AL_LEFT   = Alignment(vertical="center", horizontal="left",  wrap_text=False)
        AL_CENTER = Alignment(vertical="center", horizontal="center")

        # ── Obtener fechas con registros ─────────────────
        try:
            fechas = db.get_dias_con_registros(desde_s, hasta_s)
        except Exception:
            fechas = []

        if not fechas:
            ws = wb.active
            ws.title = "Sin datos"
            ws["A1"] = f"Sin registros entre {desde_s} y {hasta_s}"
            wb.save(path)
            self._abrir_archivo(path)
            return

        # ── Hoja por período completo ─────────────────────
        ws = wb.active
        ws.title = "Registro por día"
        ws.freeze_panes = "A3"
        ws.column_dimensions["A"].width = 5    # #
        ws.column_dimensions["B"].width = 14   # RUN
        ws.column_dimensions["C"].width = 32   # Nombre
        ws.column_dimensions["D"].width = 10   # Curso
        ws.column_dimensions["E"].width = 22   # Comidas
        ws.column_dimensions["F"].width = 10   # Total

        current_row = 1

        resumen_filas = []   # (fecha_txt, n_estudiantes, n_comidas)

        for fecha_iso in fechas:
            try:
                fecha_dt  = datetime.strptime(fecha_iso, "%Y-%m-%d")
                fecha_txt = fecha_dt.strftime("%A %d de %B de %Y").capitalize()
            except Exception:
                fecha_txt = fecha_iso

            try:
                registros_dia = db.get_asistencia_diaria(fecha_iso)
            except Exception:
                registros_dia = []

            # ── Encabezado de fecha ───────────────────────
            ws.merge_cells(
                start_row=current_row, start_column=1,
                end_row=current_row,   end_column=6
            )
            cell = ws.cell(row=current_row, column=1,
                           value=f"  {fecha_txt.upper()}")
            cell.font      = FONT_DATE
            cell.fill      = FILL_DATE
            cell.alignment = AL_LEFT
            current_row += 1

            # ── Sub-encabezado de columnas ────────────────
            for col_i, hdr in enumerate(["#", "RUN", "Apellidos, Nombres",
                                          "Curso", "Comidas", "Total"], 1):
                c = ws.cell(row=current_row, column=col_i, value=hdr)
                c.font      = FONT_HDR
                c.fill      = FILL_HDR
                c.alignment = AL_CENTER if col_i in (1, 6) else AL_LEFT
            current_row += 1

            # ── Agrupar por estudiante dentro del día ─────
            # registros_dia: lista de filas run, apellidos, nombres, curso, comida_nombre
            from collections import defaultdict
            est_comidas: dict = defaultdict(lambda: {
                "apellidos": "", "nombres": "", "curso": "", "comidas": []
            })
            for reg in registros_dia:
                run = reg.get("run") or reg.get("run_estudiante", "")
                ap  = f"{reg.get('apellido_paterno','') or ''} {reg.get('apellido_materno','') or ''}".strip()
                nom = reg.get("nombres", "") or ""
                est_comidas[run]["apellidos"] = ap
                est_comidas[run]["nombres"]   = nom
                est_comidas[run]["curso"]     = reg.get("curso", "") or ""
                est_comidas[run]["comidas"].append(reg.get("comida") or reg.get("comida_nombre", ""))

            n_est    = len(est_comidas)
            n_com    = sum(len(v["comidas"]) for v in est_comidas.values())
            resumen_filas.append((fecha_txt, n_est, n_com))

            if not est_comidas:
                c = ws.cell(row=current_row, column=1, value="(sin registros)")
                c.font = Font(italic=True, color="999999", size=10)
                ws.merge_cells(start_row=current_row, start_column=1,
                               end_row=current_row, end_column=6)
                current_row += 1
            else:
                for idx, (run, data) in enumerate(est_comidas.items(), 1):
                    nombre_full = f"{data['apellidos']}, {data['nombres']}".strip(", ")
                    comidas_txt = " · ".join(data["comidas"])
                    total_com   = len(data["comidas"])
                    run_fmt     = utils.run_display(run)

                    fill = FILL_ALT if idx % 2 == 0 else None
                    for col_i, val in enumerate(
                        [idx, run_fmt, nombre_full, data["curso"],
                         comidas_txt, total_com], 1
                    ):
                        c = ws.cell(row=current_row, column=col_i, value=val)
                        c.font = FONT_NORMAL
                        if fill:
                            c.fill = fill
                        c.alignment = AL_CENTER if col_i in (1, 6) else AL_LEFT
                    current_row += 1

            # ── Fila totales del día ──────────────────────
            ws.merge_cells(start_row=current_row, start_column=1,
                           end_row=current_row, end_column=4)
            c = ws.cell(row=current_row, column=1,
                        value=f"Total: {n_est} estudiantes")
            c.font = FONT_BOLD; c.fill = FILL_SUM; c.alignment = AL_LEFT

            ws.cell(row=current_row, column=5,
                    value=f"{n_com} comidas").font = FONT_BOLD
            ws.cell(row=current_row, column=5).fill = FILL_SUM
            ws.cell(row=current_row, column=5).alignment = AL_CENTER

            current_row += 2   # espacio entre días

        # ── Hoja resumen ──────────────────────────────────
        ws2 = wb.create_sheet("Resumen por día")
        ws2.column_dimensions["A"].width = 32
        ws2.column_dimensions["B"].width = 18
        ws2.column_dimensions["C"].width = 18

        for col_i, h in enumerate(["Fecha", "Estudiantes", "Comidas registradas"], 1):
            c = ws2.cell(row=1, column=col_i, value=h)
            c.font = FONT_HDR; c.fill = FILL_DATE; c.alignment = AL_CENTER

        for ri, (fecha_txt, n_est, n_com) in enumerate(resumen_filas, 2):
            ws2.cell(row=ri, column=1, value=fecha_txt).font = FONT_NORMAL
            ws2.cell(row=ri, column=2, value=n_est).alignment = AL_CENTER
            ws2.cell(row=ri, column=3, value=n_com).alignment = AL_CENTER

        try:
            wb.save(path)
            self._abrir_archivo(path)
        except Exception as e:
            pass

    def _export_csv_por_dia(self, desde_s: str, hasta_s: str, csv_path: str):
        """Fallback CSV si openpyxl no está disponible."""
        try:
            fechas = db.get_dias_con_registros(desde_s, hasta_s)
        except Exception:
            fechas = []
        with open(csv_path, "w", newline="", encoding="utf-8-sig") as f:
            import csv as _csv
            w = _csv.writer(f)
            for fecha_iso in fechas:
                w.writerow([f"=== {fecha_iso} ==="])
                w.writerow(["RUN", "Apellidos, Nombres", "Curso", "Comidas"])
                try:
                    regs = db.get_asistencia_diaria(fecha_iso)
                except Exception:
                    regs = []
                from collections import defaultdict
                est: dict = defaultdict(lambda: {"nombre": "", "curso": "", "comidas": []})
                for reg in regs:
                    run = reg.get("run") or reg.get("run_estudiante", "")
                    ap  = f"{reg.get('apellido_paterno','') or ''} {reg.get('apellido_materno','') or ''}".strip()
                    nom = reg.get("nombres", "") or ""
                    est[run]["nombre"] = f"{ap}, {nom}".strip(", ")
                    est[run]["curso"]  = reg.get("curso", "") or ""
                    est[run]["comidas"].append(reg.get("comida_nombre", ""))
                for run, data in est.items():
                    w.writerow([
                        utils.run_display(run), data["nombre"],
                        data["curso"], " · ".join(data["comidas"])
                    ])
                w.writerow([])
        self._abrir_archivo(csv_path)

    def _abrir_archivo(self, path: str):
        try:
            import platform
            if platform.system() == "Darwin":
                subprocess.Popen(["open", path])
            elif platform.system() == "Windows":
                os.startfile(path)
        except Exception:
            pass


_TBL_STYLE = f"""
    QTableWidget {{
        background: {C.SURFACE}; border: none;
        border-radius: 10px; outline: none;
    }}
    QTableWidget::item {{ padding: 6px 12px; border: none; }}
    QTableWidget::item:selected {{ background: {C.BLUE_DIM}; color: {C.TEXT}; }}
    QTableWidget::item:alternate {{ background: {C.SURFACE2}; }}
    QHeaderView::section {{
        background: {C.SURFACE2}; color: {C.TEXT2};
        padding: 8px 12px; border: none; font-size: 11px; font-weight: 600;
    }}
    QHeaderView::section:hover {{ background: {C.SURFACE3}; cursor: pointer; }}
"""

_TIPO_COLOR_MAP = {
    "atraso":       C.AMBER,
    "inasistencia": C.RED,
    "retiro":       C.BLUE,
}


def _periodo_filter_btn(label: str, active: bool = False) -> QPushButton:
    btn = QPushButton(label)
    btn.setCheckable(True)
    btn.setChecked(active)
    btn.setCursor(Qt.CursorShape.PointingHandCursor)
    btn.setFixedHeight(30)
    _apply_filter_style(btn)
    return btn


def _apply_filter_style(btn: QPushButton):
    if btn.isChecked():
        btn.setStyleSheet(f"""
            QPushButton {{
                background: {C.BLUE}; color: white;
                border: none;
                border-bottom: 3px solid #005BBB;
                border-radius: 8px 8px 0 0;
                padding: 0 16px; font-size: 12px; font-weight: 600;
            }}
        """)
    else:
        btn.setStyleSheet(f"""
            QPushButton {{
                background: transparent; color: {C.TEXT2};
                border: 1px solid {C.BORDER}; border-radius: 8px;
                padding: 0 16px; font-size: 12px;
            }}
            QPushButton:hover {{ background: {C.SURFACE2}; color: {C.TEXT}; }}
        """)


class _InspectoriaTab(QWidget):
    """
    Reportes de Inspectoría — #18:
    ┌ KPIs: Atrasos / Inasistencias / Retiros / Sin firma (mes) ┐
    ├ Sub-tab "Registro de Pases"                                │
    │   Filtros: Hoy / Semana / Mes   [Exportar CSV]            │
    │   Tabla: Fecha | Hora | RUN | Nombre | Curso | Tipo | Firma│
    └ Sub-tab "Top Mensual"                                      ┘
        Columnas: # | Nombre | Atrasos | Inasistencias | Retiros | Total
        Ordenable por cualquier columna numérica
    """

    # Índices de columnas numéricas en el top
    _COL_ATRASOS     = 4
    _COL_INASISTENCIAS = 5
    _COL_RETIROS     = 6
    _COL_TOTAL       = 7
    _SORT_DEFAULT    = _COL_TOTAL

    def __init__(self, parent=None):
        super().__init__(parent)
        self._sort_col     = self._SORT_DEFAULT
        self._periodo      = "mes"       # "hoy" | "semana" | "mes"
        self._pases_rows   = []          # cache para export
        self._build_ui()
        self.refresh()

    def _build_ui(self):
        root = QVBoxLayout(self)
        root.setContentsMargins(16, 16, 16, 16)
        root.setSpacing(12)

        # ── KPI cards ─────────────────────────────────────────
        kpi_row = QHBoxLayout()
        kpi_row.setSpacing(10)
        self._kpi_atrasos  = StatCard("Atrasos mes",       "—", accent=C.AMBER)
        self._kpi_inasis   = StatCard("Inasistencias mes", "—", accent=C.RED)
        self._kpi_retiros  = StatCard("Retiros mes",       "—", accent=C.BLUE)
        self._kpi_sinfirma = StatCard("Sin firma",         "—", accent="#FF6B00")
        self._kpi_wa       = StatCard("WhatsApp enviados", "—", accent="#25D366")
        for card in (self._kpi_atrasos, self._kpi_inasis, self._kpi_retiros,
                     self._kpi_sinfirma, self._kpi_wa):
            card.setMinimumHeight(72)
            kpi_row.addWidget(card)
        root.addLayout(kpi_row)

        # ── Sub-tabs: Registro de Pases / Top Mensual ─────────
        self._sub_tabs = QTabWidget()
        self._sub_tabs.setStyleSheet(f"""
            QTabWidget::pane {{ border: none; background: {C.BG}; }}
            QTabBar::tab {{
                background: transparent; color: {C.TEXT3};
                border: none; padding: 8px 18px; font-size: 12px;
                font-weight: 500;
            }}
            QTabBar::tab:selected {{
                color: {C.TEXT}; font-weight: 700;
                border-bottom: 2px solid {C.BLUE};
            }}
            QTabBar::tab:hover:!selected {{ color: {C.TEXT2}; }}
        """)

        self._sub_tabs.addTab(self._build_registro_tab(), "Registro de pases")
        self._sub_tabs.addTab(self._build_top_tab(),      "Top mensual")
        root.addWidget(self._sub_tabs, stretch=1)

    # ── Sub-tab: Registro de Pases ────────────────────────────

    def _build_registro_tab(self) -> QWidget:
        w = QWidget()
        lay = QVBoxLayout(w)
        lay.setContentsMargins(0, 12, 0, 0)
        lay.setSpacing(10)

        # Fila de filtros + export
        ctrl = QHBoxLayout()
        ctrl.setSpacing(6)

        self._btn_hoy    = _periodo_filter_btn("Hoy")
        self._btn_semana = _periodo_filter_btn("Semana")
        self._btn_mes    = _periodo_filter_btn("Mes", active=True)

        for btn, p in ((self._btn_hoy, "hoy"), (self._btn_semana, "semana"), (self._btn_mes, "mes")):
            btn.clicked.connect(lambda _, period=p: self._set_periodo(period))
            ctrl.addWidget(btn)

        ctrl.addStretch()

        self._lbl_count = QLabel("0 pases")
        self._lbl_count.setStyleSheet(
            f"font-size: 11px; color: {C.TEXT3}; background: transparent;"
        )
        ctrl.addWidget(self._lbl_count)

        btn_export = AButton("⬇  Exportar CSV", sound_type="click")
        btn_export.setStyleSheet(f"""
            QPushButton {{
                background: transparent; color: {C.BLUE};
                border: 1px solid {C.BLUE}44; border-radius: 8px;
                padding: 5px 14px; font-size: 12px;
            }}
            QPushButton:hover {{ background: {C.BLUE_DIM}; }}
        """)
        btn_export.clicked.connect(self._export_csv)
        ctrl.addWidget(btn_export)

        lay.addLayout(ctrl)

        # Tabla de pases
        self._tbl_pases = QTableWidget()
        self._tbl_pases.setColumnCount(8)
        self._tbl_pases.setHorizontalHeaderLabels(
            ["Fecha", "Hora", "RUN", "Nombre", "Curso", "Tipo", "Firmado", "Firmado por"]
        )
        self._tbl_pases.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        self._tbl_pases.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._tbl_pases.verticalHeader().setVisible(False)
        self._tbl_pases.setAlternatingRowColors(True)
        self._tbl_pases.setShowGrid(False)
        self._tbl_pases.verticalHeader().setDefaultSectionSize(34)
        self._tbl_pases.setStyleSheet(_TBL_STYLE)
        lay.addWidget(self._tbl_pases, stretch=1)

        return w

    # ── Sub-tab: Top Mensual ──────────────────────────────────

    def _build_top_tab(self) -> QWidget:
        w = QWidget()
        lay = QVBoxLayout(w)
        lay.setContentsMargins(0, 12, 0, 0)
        lay.setSpacing(10)

        # Botones de ordenación
        sort_row = QHBoxLayout()
        sort_row.setSpacing(6)
        lbl_sort = QLabel("Ordenar por:")
        lbl_sort.setStyleSheet(f"font-size: 11px; color: {C.TEXT3}; background: transparent;")
        sort_row.addWidget(lbl_sort)

        self._sort_btns: dict[int, QPushButton] = {}
        for col, label in (
            (self._COL_TOTAL,       "Total"),
            (self._COL_ATRASOS,     "Atrasos"),
            (self._COL_INASISTENCIAS, "Inasistencias"),
            (self._COL_RETIROS,     "Retiros"),
        ):
            btn = _periodo_filter_btn(label, active=(col == self._sort_col))
            btn.clicked.connect(lambda _, c=col: self._sort_top(c))
            sort_row.addWidget(btn)
            self._sort_btns[col] = btn

        sort_row.addStretch()

        btn_export_top = AButton("⬇  Exportar CSV", sound_type="click")
        btn_export_top.setStyleSheet(f"""
            QPushButton {{
                background: transparent; color: {C.BLUE};
                border: 1px solid {C.BLUE}44; border-radius: 8px;
                padding: 5px 14px; font-size: 12px;
            }}
            QPushButton:hover {{ background: {C.BLUE_DIM}; }}
        """)
        btn_export_top.clicked.connect(self._export_top_csv)
        sort_row.addWidget(btn_export_top)

        lay.addLayout(sort_row)

        # Tabla top
        self._tbl_top = QTableWidget()
        self._tbl_top.setColumnCount(8)
        self._tbl_top.setHorizontalHeaderLabels(
            ["#", "RUN", "Nombre", "Curso", "Atrasos", "Inasistencias", "Retiros", "Total"]
        )
        self._tbl_top.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        self._tbl_top.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        self._tbl_top.verticalHeader().setVisible(False)
        self._tbl_top.setAlternatingRowColors(True)
        self._tbl_top.setShowGrid(False)
        self._tbl_top.verticalHeader().setDefaultSectionSize(34)
        self._tbl_top.setStyleSheet(_TBL_STYLE)
        # Clic en header → ordenar
        self._tbl_top.horizontalHeader().sectionClicked.connect(self._on_top_header_click)
        lay.addWidget(self._tbl_top, stretch=1)

        return w

    # ── Lógica de período ─────────────────────────────────────

    def _set_periodo(self, periodo: str):
        self._periodo = periodo
        for btn, p in ((self._btn_hoy, "hoy"), (self._btn_semana, "semana"), (self._btn_mes, "mes")):
            btn.setChecked(p == periodo)
            _apply_filter_style(btn)
        self._refresh_registro()

    def _get_rango(self):
        today = date.today()
        if self._periodo == "hoy":
            return today.isoformat(), today.isoformat()
        elif self._periodo == "semana":
            lunes = today - timedelta(days=today.weekday())
            return lunes.isoformat(), today.isoformat()
        else:  # mes
            return today.replace(day=1).isoformat(), today.isoformat()

    # ── Refresh principal ─────────────────────────────────────

    def refresh(self):
        today      = date.today()
        inicio_mes = today.replace(day=1).isoformat()

        conn = db.get_conn()

        def _count(tipo):
            return conn.execute(
                "SELECT COUNT(*) AS n FROM student_suspensions WHERE tipo=? AND fecha_inicio >= ?",
                (tipo, inicio_mes)
            ).fetchone()["n"]

        n_atr = _count("atraso")
        n_ina = _count("inasistencia")
        n_ret = _count("retiro")
        n_sf  = conn.execute(
            "SELECT COUNT(*) AS n FROM student_suspensions "
            "WHERE tipo IN ('atraso','inasistencia','retiro') "
            "AND firmado=0 AND fecha_inicio >= ?",
            (inicio_mes,)
        ).fetchone()["n"]

        try:
            stats = db.get_whatsapp_stats(desde=inicio_mes)
            n_wa = stats.get("enviados", 0)
        except Exception:
            n_wa = 0

        conn.close()

        self._kpi_atrasos.set_value(str(n_atr))
        self._kpi_inasis.set_value(str(n_ina))
        self._kpi_retiros.set_value(str(n_ret))
        self._kpi_sinfirma.set_value(str(n_sf))
        self._kpi_wa.set_value(str(n_wa))

        self._refresh_registro()
        self._refresh_top()

    def _refresh_registro(self):
        import session as _sess
        desde, hasta = self._get_rango()
        try:
            rows = db.get_pases_periodo(desde, hasta,
                                        periodo=_sess.viewing_period())
        except Exception:
            rows = []

        self._pases_rows = rows
        self._lbl_count.setText(f"{len(rows)} pase{'s' if len(rows) != 1 else ''}")

        self._tbl_pases.setRowCount(len(rows))
        for i, r in enumerate(rows):
            # Hora aproximada desde creado_en
            creado = r["creado_en"] or ""
            hora   = creado[11:16] if len(creado) >= 16 else "—"
            fecha  = r["fecha_inicio"] or ""
            try:
                fecha = date.fromisoformat(fecha).strftime("%d/%m/%Y")
            except Exception:
                pass
            ap    = f"{r['apellido_paterno'] or ''} {r['apellido_materno'] or ''}".strip()
            nombre = f"{ap}, {r['nombres'] or ''}".strip(", ")
            curso  = db.display_curso(r["curso"] or "", db.get_cursos_nombres_map()) or "—"
            tipo   = r["tipo"] or ""
            firmado = "✓" if r["firmado"] else "✗"
            firmado_por = r["firmado_por"] or "—"

            vals = [fecha, hora, utils.run_display(r["run"]), nombre,
                    curso, tipo.capitalize(), firmado, firmado_por]
            for j, v in enumerate(vals):
                it = QTableWidgetItem(str(v))
                if j == 5:   # tipo
                    it.setForeground(QColor(_TIPO_COLOR_MAP.get(tipo, C.TEXT2)))
                elif j == 6:  # firmado
                    it.setForeground(QColor(C.GREEN if r["firmado"] else C.RED))
                    it.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                self._tbl_pases.setItem(i, j, it)

        for col in (0, 1, 2, 5, 6):
            self._tbl_pases.resizeColumnToContents(col)

    def _refresh_top(self):
        try:
            rows = db.get_top_pases_mes(limit=50)
        except Exception:
            rows = []

        # rows son dicts — ordenar por clave de texto
        col_key_map = {
            self._COL_ATRASOS:       "atrasos",
            self._COL_INASISTENCIAS: "inasistencias",
            self._COL_RETIROS:       "retiros",
            self._COL_TOTAL:         "total",
        }
        sort_key = col_key_map.get(self._sort_col, "total")
        try:
            rows = sorted(rows, key=lambda r: int(r[sort_key] or 0), reverse=True)
        except Exception:
            pass

        self._tbl_top.setRowCount(len(rows))
        for i, r in enumerate(rows):
            run    = r["run"]
            ap     = f"{r.get('apellido_paterno') or ''} {r.get('apellido_materno') or ''}".strip()
            nombre = f"{ap}, {r.get('nombres') or ''}".strip(", ")
            curso  = db.display_curso(r.get("curso") or "", db.get_cursos_nombres_map()) or "—"
            atr    = int(r.get("atrasos", 0) or 0)
            ina    = int(r.get("inasistencias", 0) or 0)
            ret    = int(r.get("retiros", 0) or 0)
            total  = int(r.get("total", 0) or 0)
            num_vals = (atr, ina, ret, total)

            row_data = [str(i + 1), utils.run_display(run), nombre, curso,
                        str(atr), str(ina), str(ret), str(total)]
            for j, v in enumerate(row_data):
                it = QTableWidgetItem(v)
                if j >= 4:
                    it.setTextAlignment(Qt.AlignmentFlag.AlignCenter)
                    num = num_vals[j - 4]
                    if num >= 5:
                        it.setForeground(QColor(C.RED))
                    elif num >= 3:
                        it.setForeground(QColor(C.AMBER))
                self._tbl_top.setItem(i, j, it)

        for col in (0, 1, 4, 5, 6, 7):
            self._tbl_top.resizeColumnToContents(col)

    # ── Ordenación ────────────────────────────────────────────

    def _on_top_header_click(self, logical_index: int):
        if logical_index in (self._COL_ATRASOS, self._COL_INASISTENCIAS,
                              self._COL_RETIROS, self._COL_TOTAL):
            self._sort_top(logical_index)

    def _sort_top(self, col: int):
        self._sort_col = col
        for c, btn in self._sort_btns.items():
            btn.setChecked(c == col)
            _apply_filter_style(btn)
        self._refresh_top()

    # ── Exportar CSV ──────────────────────────────────────────

    def _export_csv(self):
        if not self._pases_rows:
            return
        desde, hasta = self._get_rango()
        filename = f"pases_{self._periodo}_{date.today().isoformat()}.csv"
        path = os.path.join(os.path.expanduser("~/Desktop"), filename)
        try:
            with open(path, "w", newline="", encoding="utf-8-sig") as f:
                w = csv.writer(f)
                w.writerow(["Fecha", "Hora", "RUN", "Nombre", "Apellido P.",
                             "Apellido M.", "Curso", "Tipo", "Firmado", "Firmado por"])
                for r in self._pases_rows:
                    creado = r["creado_en"] or ""
                    hora   = creado[11:16] if len(creado) >= 16 else ""
                    w.writerow([
                        r["fecha_inicio"],
                        hora,
                        utils.run_display(r["run"]),
                        r["nombres"] or "",
                        r["apellido_paterno"] or "",
                        r["apellido_materno"] or "",
                        r["curso"] or "",
                        r["tipo"] or "",
                        "Sí" if r["firmado"] else "No",
                        r["firmado_por"] or "",
                    ])
            if os.path.exists(path):
                subprocess.Popen(["open", os.path.dirname(path)])
        except Exception:
            pass

    def _export_top_csv(self):
        try:
            rows = db.get_top_pases_mes(limit=50)
        except Exception:
            rows = []
        if not rows:
            return
        filename = f"top_pases_{date.today().isoformat()}.csv"
        path = os.path.join(os.path.expanduser("~/Desktop"), filename)
        try:
            with open(path, "w", newline="", encoding="utf-8-sig") as f:
                w = csv.writer(f)
                w.writerow(["#", "RUN", "Nombre", "Apellido P.", "Apellido M.",
                             "Curso", "Atrasos", "Inasistencias", "Retiros", "Total"])
                for i, r in enumerate(rows, 1):
                    run, nombres, ap_pat, ap_mat, curso, atr, ina, ret, total = r
                    w.writerow([i, utils.run_display(run), nombres or "",
                                 ap_pat or "", ap_mat or "", curso or "",
                                 atr, ina, ret, total])
            if os.path.exists(path):
                subprocess.Popen(["open", os.path.dirname(path)])
        except Exception:
            pass


class ReportsScreen(QWidget):

    def __init__(self, parent=None):
        super().__init__(parent)
        # Inicializar referencias de acción (se asignan dentro de _wrap_with_action)
        self._action_mes:    tuple | None = None
        self._action_espera: tuple | None = None
        self._build_ui()
        # Overlays flotantes (deben ir después de _build_ui para que self tenga tamaño)
        self._toast   = ToastBanner(self)
        self._confirm = ConfirmPanel(self)
        self._load_data()

    def _build_ui(self):
        self.setStyleSheet(f"background: {C.BG};")
        root = QVBoxLayout(self)
        root.setContentsMargins(24, 20, 24, 20)
        root.setSpacing(16)

        # ── Header ──────────────────────────────────
        hdr = QHBoxLayout()

        title = QLabel("Reportes")
        title.setStyleSheet(
            f"font-size: 22px; font-weight: 700; color: {C.TEXT}; background: transparent;"
        )
        hdr.addWidget(title)
        hdr.addStretch()

        self._lbl_range = QLabel("")
        self._lbl_range.setStyleSheet(
            f"font-size: 12px; color: {C.TEXT3}; background: transparent;"
        )
        hdr.addWidget(self._lbl_range)
        hdr.addSpacing(12)

        _btn_style = f"""
            QPushButton {{
                background: transparent;
                color: {C.TEXT2};
                border: 1.5px solid {C.BORDER};
                border-radius: 8px;
                padding: 7px 14px;
                font-size: 12px;
                font-weight: 600;
            }}
            QPushButton:hover {{ background: {C.SURFACE2}; color: {C.TEXT}; }}
        """

        btn_export = AButton("↓  Exportar CSV", sound_type="click")
        btn_export.setStyleSheet(_btn_style)
        btn_export.clicked.connect(self._export_current_tab)
        hdr.addWidget(btn_export)
        hdr.addSpacing(6)

        btn_refresh = AButton("↻  Actualizar", sound_type="click")
        btn_refresh.setStyleSheet(_btn_style)
        btn_refresh.clicked.connect(self._load_data)
        hdr.addWidget(btn_refresh)
        root.addLayout(hdr)

        # ── KPI row ──────────────────────────────────
        kpi_row = QHBoxLayout()
        kpi_row.setSpacing(12)

        self._kpi_regs    = StatCard("Registros semana",  "—", accent=C.NAVY_400)
        self._kpi_strikes = StatCard("Strikes semana",    "—", accent=C.RED)
        self._kpi_afect   = StatCard("Con strikes",       "—", accent=C.AMBER)

        for card in (self._kpi_regs, self._kpi_strikes, self._kpi_afect):
            card.setMinimumHeight(88)
            kpi_row.addWidget(card)

        root.addLayout(kpi_row)

        # ── Tabs ─────────────────────────────────────
        self._tabs = QTabWidget()
        self._tabs.setStyleSheet(f"""
            QTabWidget::pane {{
                border: 1.5px solid {C.BORDER};
                border-radius: 12px;
                background: {C.SURFACE};
                top: -1px;
            }}
            QTabBar::tab {{
                background: transparent;
                color: {C.TEXT2};
                padding: 9px 22px;
                border: none;
                border-radius: 8px;
                margin: 2px;
                font-weight: 500;
                font-size: 13px;
            }}
            QTabBar::tab:selected {{
                background: {C.BLUE_DIM};
                color: {C.BLUE};
                font-weight: 700;
                border-bottom: 2px solid {C.BLUE};
            }}
            QTabBar::tab:hover:!selected {{
                background: {C.SURFACE};
                color: {C.TEXT};
            }}
        """)

        # ── Tabs con contenedores (tabla + barra de acción) ──
        import session as _sess
        _rol = _sess.rol()

        # Crear widgets solo para los roles que los usan
        if _rol in ("admin", "pae"):
            self._tbl_mes   = self._make_table()
            self._tab_asist = _AsistenciaTab()
            self._tbl_log   = self._make_log_table()
            self._tbl_mes.itemSelectionChanged.connect(
                lambda: self._on_sel_changed(self._tbl_mes, self._action_mes))
            self._tabs.addTab(self._wrap_with_action(
                self._tbl_mes, "baja"),        "Top 25 — Mes")
            self._tabs.addTab(self._tab_asist, "Asistencia")
            self._tabs.addTab(self._tbl_log,   "Log altas/bajas")

        if _rol == "admin":
            self._tbl_espera = self._make_espera_table()
            self._tbl_espera.itemSelectionChanged.connect(
                lambda: self._on_sel_changed(self._tbl_espera, self._action_espera))
            self._tabs.addTab(self._wrap_with_action(
                self._tbl_espera, "promover"), "Lista de espera")

        if _rol in ("admin", "inspectoria"):
            self._tab_insp = _InspectoriaTab()
            self._tabs.addTab(self._tab_insp, "Inspectoría")

        root.addWidget(self._tabs, stretch=1)

    def _make_table(self) -> QTableWidget:
        t = QTableWidget()
        t.setColumnCount(6)
        t.setHorizontalHeaderLabels(["#", "RUN", "Nombre", "Curso", "Strikes", "Riesgo"])
        t.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        t.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        t.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        t.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        t.verticalHeader().setVisible(False)
        t.setAlternatingRowColors(True)
        t.setShowGrid(False)
        t.verticalHeader().setDefaultSectionSize(40)
        t.setStyleSheet(f"""
            QTableWidget {{
                background: {C.SURFACE};
                border: none;
                border-radius: 12px;
                outline: none;
            }}
            QTableWidget::item {{
                padding: 8px 12px;
                border: none;
            }}
            QTableWidget::item:selected {{
                background: {C.NAVY_700};
                color: {C.TEXT};
            }}
            QTableWidget::item:alternate {{
                background: {C.SURFACE2};
            }}
        """)
        return t

    def _make_espera_table(self) -> QTableWidget:
        t = QTableWidget()
        t.setColumnCount(6)
        t.setHorizontalHeaderLabels(["#", "RUN", "Apellidos", "Nombres", "Curso", "RSH %"])
        t.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        t.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeMode.Stretch)
        t.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        t.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        t.verticalHeader().setVisible(False)
        t.setAlternatingRowColors(True)
        t.setShowGrid(False)
        t.verticalHeader().setDefaultSectionSize(40)
        t.setStyleSheet(self._make_table().styleSheet())
        return t

    def _make_log_table(self) -> QTableWidget:
        t = QTableWidget()
        t.setColumnCount(6)
        t.setHorizontalHeaderLabels(["Fecha/Hora", "RUN", "Nombre", "Curso", "Antes", "Ahora"])
        t.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeMode.Stretch)
        t.setEditTriggers(QAbstractItemView.EditTrigger.NoEditTriggers)
        t.setSelectionBehavior(QAbstractItemView.SelectionBehavior.SelectRows)
        t.verticalHeader().setVisible(False)
        t.setAlternatingRowColors(True)
        t.setShowGrid(False)
        t.verticalHeader().setDefaultSectionSize(40)
        t.setStyleSheet(self._make_table().styleSheet())
        return t

    def _load_data(self):
        import session as _sess
        _rol = _sess.rol()

        if _rol in ("admin", "pae"):
            res = db.get_resumen_semana()
            self._kpi_regs.set_value(str(res.get("total_registros", 0)))
            self._kpi_regs.flash()
            self._kpi_strikes.set_value(str(res.get("total_strikes", 0)))
            self._kpi_strikes.flash(C.RED)
            self._kpi_afect.set_value(str(res.get("estudiantes_con_strikes", 0)))
            self._kpi_afect.flash(C.AMBER)

            desde = utils.format_fecha_display(res["desde"])
            hasta = utils.format_fecha_display(res["hasta"])
            self._lbl_range.setText(f"{desde}  →  {hasta}")

            self._fill_table(self._tbl_mes, db.get_top_ausentes_mes(25))
            self._tab_asist._load()
            self._fill_log(db.get_status_log(200))

        if _rol == "admin":
            self._fill_espera(db.get_waitlist_sorted())

        if _rol in ("admin", "inspectoria"):
            self._tab_insp.refresh()

    def _fill_table(self, tabla: QTableWidget, rows: list):
        tabla.setRowCount(len(rows))

        for i, row in enumerate(rows):
            rank    = i + 1
            color   = _rank_color(rank)
            run_fmt = utils.run_display(row["run_estudiante"])
            nombre  = (
                f"{row.get('apellido_paterno','') or ''} "
                f"{row.get('apellido_materno','') or ''}, "
                f"{row.get('nombres','') or ''}"
            ).strip(", ")
            curso   = row.get("curso", "") or ""
            strikes = row["total_strikes"]

            # Risk label
            if rank <= 5:    riesgo = "Crítico"
            elif rank <= 10: riesgo = "Alto"
            elif rank <= 15: riesgo = "Medio"
            else:            riesgo = "Bajo"

            values = [str(rank), run_fmt, nombre, curso, str(strikes), riesgo]

            for col, txt in enumerate(values):
                item = QTableWidgetItem(txt)
                item.setTextAlignment(
                    Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft
                )
                # Guardar RUN raw para recuperarlo en operaciones masivas
                if col == 1:
                    item.setData(Qt.ItemDataRole.UserRole, row["run_estudiante"])

                if col == 0:   # Rank number
                    item.setForeground(QColor(color))
                    f = item.font(); f.setBold(True); item.setFont(f)
                    item.setTextAlignment(
                        Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignCenter
                    )
                elif col == 4:  # Strikes
                    item.setForeground(QColor(color))
                    f = item.font(); f.setBold(True); item.setFont(f)
                elif col == 5:  # Risk
                    item.setForeground(QColor(color))

                tabla.setItem(i, col, item)

        tabla.resizeRowsToContents()
        # Copiar celda al hacer click (col 1 = RUN, col 2 = Nombre)
        make_table_copyable(tabla, copy_cols=[1, 2])

    def _fill_espera(self, rows: list):
        """Llena la tab Lista de espera con estudiantes priorizados por RSH."""
        _ESTADO = {
            "beneficiario":    ("Beneficiario PAE", C.GREEN),
            "espera":          ("Lista de espera",  C.AMBER),
            "no_beneficiario": ("No beneficiario",  C.TEXT3),
        }
        self._tbl_espera.setRowCount(len(rows))
        for i, r in enumerate(rows):
            run_fmt   = utils.run_display(r["run"])
            apellidos = f"{r.get('apellido_paterno','') or ''} {r.get('apellido_materno','') or ''}".strip()
            nombres   = r.get("nombres", "") or ""
            curso     = r.get("curso", "") or ""
            rsh       = r.get("puntaje_rsh")
            rsh_txt   = str(rsh) if rsh is not None else "—"

            for col, txt in enumerate([str(i + 1), run_fmt, apellidos, nombres, curso, rsh_txt]):
                item = QTableWidgetItem(txt)
                item.setTextAlignment(Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft)
                # RUN raw en UserRole para operaciones masivas
                if col == 1:
                    item.setData(Qt.ItemDataRole.UserRole, r["run"])
                if col == 5 and rsh is not None:
                    item.setForeground(QColor(
                        C.RED if rsh <= 40 else (C.AMBER if rsh <= 70 else C.TEXT2)
                    ))
                self._tbl_espera.setItem(i, col, item)
        make_table_copyable(self._tbl_espera, copy_cols=[1, 2, 3])

    def _fill_log(self, rows: list):
        """Llena la tab Log altas/bajas."""
        _LABELS = {
            "beneficiario":    ("Beneficiario PAE", C.GREEN),
            "espera":          ("Lista de espera",  C.AMBER),
            "no_beneficiario": ("No beneficiario",  C.TEXT3),
        }
        self._tbl_log.setRowCount(len(rows))
        for i, r in enumerate(rows):
            ts_raw = r.get("timestamp", "")
            try:
                ts_fmt = datetime.fromisoformat(ts_raw).strftime("%d/%m/%Y %H:%M")
            except Exception:
                ts_fmt = ts_raw
            run_fmt   = utils.run_display(r["run"])
            ap        = r.get("apellido_paterno", "") or ""
            nom       = r.get("nombres", "") or ""
            nombre    = f"{ap}, {nom}".strip(", ")
            curso     = r.get("curso", "") or ""
            ant_lbl, ant_clr = _LABELS.get(r["estado_ant"], (r["estado_ant"], C.TEXT3))
            new_lbl, new_clr = _LABELS.get(r["estado_new"], (r["estado_new"], C.TEXT3))

            values = [ts_fmt, run_fmt, nombre, curso, ant_lbl, new_lbl]
            for col, txt in enumerate(values):
                item = QTableWidgetItem(txt)
                item.setTextAlignment(Qt.AlignmentFlag.AlignVCenter | Qt.AlignmentFlag.AlignLeft)
                if col == 4:
                    item.setForeground(QColor(ant_clr))
                elif col == 5:
                    item.setForeground(QColor(new_clr))
                    f = item.font(); f.setBold(True); item.setFont(f)
                self._tbl_log.setItem(i, col, item)

    def _export_current_tab(self):
        """Exporta la tabla del tab activo. Asistencia → Excel, resto → CSV."""
        import session as _sess
        _rol    = _sess.rol()
        tab_idx = self._tabs.currentIndex()
        tab_widget = self._tabs.widget(tab_idx)

        # Asistencia → exportador propio (Excel)
        if tab_widget is self._tab_asist if hasattr(self, "_tab_asist") else False:
            self._tab_asist._export_excel()
            return

        # Inspectoría → exportador propio (CSV)
        if hasattr(self, "_tab_insp") and tab_widget is self._tab_insp:
            self._tab_insp._export_csv()
            return

        # Buscar la tabla dentro del widget actual (puede estar en un _wrap_with_action)
        tabla: QTableWidget | None = None
        tab_name = "reporte"
        for attr, name in [
            ("_tbl_mes",    "top25_mes"),
            ("_tbl_espera", "lista_espera"),
            ("_tbl_log",    "log_altas_bajas"),
        ]:
            tbl = getattr(self, attr, None)
            if tbl is None:
                continue
            # El tab puede ser el QTableWidget mismo o un wrapper con la tabla adentro
            if tab_widget is tbl or tbl.parent() is tab_widget or tbl.parent() is tab_widget.parent():
                tabla    = tbl
                tab_name = name
                break

        if tabla is None:
            return

        export_dir = os.path.join(os.path.expanduser("~"), "pae_control", "exports")
        os.makedirs(export_dir, exist_ok=True)

        ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
        path = os.path.join(export_dir, f"{tab_name}_{ts}.csv")

        headers = [
            tabla.horizontalHeaderItem(c).text()
            for c in range(tabla.columnCount())
        ]

        with open(path, "w", newline="", encoding="utf-8-sig") as f:
            writer = csv.writer(f)
            writer.writerow(headers)
            for row in range(tabla.rowCount()):
                writer.writerow([
                    (tabla.item(row, col).text() if tabla.item(row, col) else "")
                    for col in range(tabla.columnCount())
                ])

        try:
            subprocess.Popen(["open", path])
        except Exception:
            pass

    # ─────────────────────────────────────────────
    #  BARRAS DE ACCIÓN MASIVA
    # ─────────────────────────────────────────────

    def _wrap_with_action(self, tabla: QTableWidget, modo: str) -> QWidget:
        """
        Envuelve una tabla en un contenedor QWidget con una barra de
        acción en la parte inferior (aparece solo cuando hay selección).
        modo: "baja" | "promover"
        """
        container = QWidget()
        container.setStyleSheet(f"background: {C.BG};")
        lay = QVBoxLayout(container)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(0)
        lay.addWidget(tabla, stretch=1)

        bar = QFrame()
        bar.setFixedHeight(52)
        bar.setStyleSheet(f"""
            QFrame {{
                background: {C.SURFACE};
                border-top: 1px solid {C.BORDER};
            }}
        """)
        bar_lay = QHBoxLayout(bar)
        bar_lay.setContentsMargins(16, 0, 16, 0)
        bar_lay.setSpacing(12)

        lbl = QLabel("")
        lbl.setStyleSheet(f"font-size: 12px; color: {C.TEXT2}; background: transparent;")
        bar_lay.addWidget(lbl)
        bar_lay.addStretch()

        if modo == "baja":
            btn = AButton("↓  Dar de baja seleccionados", sound_type="click")
            btn.setStyleSheet(f"""
                QPushButton {{
                    background: {C.RED_DIM}; color: {C.RED};
                    border: 1.5px solid {C.RED}55; border-radius: 8px;
                    padding: 0 18px; font-size: 12px; font-weight: 700;
                }}
                QPushButton:hover {{ background: {C.RED}22; }}
                QPushButton:disabled {{ background: {C.SURFACE2}; color: {C.TEXT3}; border-color: transparent; }}
            """)
            btn.clicked.connect(lambda: self._pedir_baja(tabla))
        else:
            btn = AButton("↑  Promover al PAE", sound_type="click")
            btn.setStyleSheet(f"""
                QPushButton {{
                    background: {C.GREEN_DIM}; color: {C.GREEN};
                    border: 1.5px solid {C.GREEN}55; border-radius: 8px;
                    padding: 0 18px; font-size: 12px; font-weight: 700;
                }}
                QPushButton:hover {{ background: {C.GREEN}22; }}
                QPushButton:disabled {{ background: {C.SURFACE2}; color: {C.TEXT3}; border-color: transparent; }}
            """)
            btn.clicked.connect(lambda: self._pedir_promover(tabla))

        btn.setEnabled(False)
        bar_lay.addWidget(btn)
        lay.addWidget(bar)

        # Guardar referencias por tabla
        if modo == "baja" and tabla is self._tbl_mes:
            self._action_mes    = (lbl, btn, bar)
        elif modo == "promover":
            self._action_espera = (lbl, btn, bar)

        bar.hide()
        return container

    def _on_sel_changed(self, tabla: QTableWidget, action_tuple):
        lbl, btn, bar = action_tuple
        rows = tabla.selectionModel().selectedRows()
        n    = len(rows)
        if n > 0:
            lbl.setText(f"{n} estudiante{'s' if n > 1 else ''} seleccionado{'s' if n > 1 else ''}")
            btn.setEnabled(True)
            bar.show()
        else:
            btn.setEnabled(False)
            bar.hide()

    def _get_selected_items(self, tabla: QTableWidget) -> list[dict]:
        """Retorna [{run, nombre}] de las filas seleccionadas."""
        items = []
        seen  = set()
        for idx in tabla.selectionModel().selectedRows():
            row = idx.row()
            run_item = tabla.item(row, 1)
            nom_item = tabla.item(row, 2)
            if run_item:
                run_raw = run_item.data(Qt.ItemDataRole.UserRole) or run_item.text()
                if run_raw not in seen:
                    seen.add(run_raw)
                    items.append({
                        "run":    run_raw,
                        "nombre": nom_item.text() if nom_item else run_raw,
                    })
        return items

    # ─────────────────────────────────────────────
    #  OPERACIONES MASIVAS
    # ─────────────────────────────────────────────

    def _pedir_baja(self, tabla: QTableWidget):
        items = self._get_selected_items(tabla)
        if not items:
            return
        n = len(items)
        self._confirm.pedir_confirmacion(
            accion=f"Dar de baja a {n} estudiante{'s' if n > 1 else ''}  —  Esta acción los desactiva del PAE",
            color=C.RED,
            items=items,
            on_ok=self._ejecutar_baja,
        )

    def _ejecutar_baja(self, runs: list[str]):
        ok = 0
        for run in runs:
            try:
                db.update_student_status(run, activo=0, lista_espera=0,
                                         motivo="Baja manual desde reportes")
                ok += 1
            except Exception:
                pass
        sound.save()
        self._toast.show_toast(
            f"✓  {ok} estudiante{'s' if ok > 1 else ''} dado{'s' if ok > 1 else ''} de baja",
            tipo="ok",
        )
        self._load_data()

    def _pedir_promover(self, tabla: QTableWidget):
        items = self._get_selected_items(tabla)
        if not items:
            return
        n = len(items)
        self._confirm.pedir_confirmacion(
            accion=f"Promover {n} estudiante{'s' if n > 1 else ''} al PAE  —  Pasarán de lista de espera a beneficiarios",
            color=C.GREEN,
            items=items,
            on_ok=self._ejecutar_promover,
        )

    def _ejecutar_promover(self, runs: list[str]):
        ok = 0
        for run in runs:
            try:
                db.update_student_status(run, activo=1, lista_espera=0,
                                         motivo="Promoción manual desde reportes")
                ok += 1
            except Exception:
                pass
        sound.save()
        self._toast.show_toast(
            f"✓  {ok} estudiante{'s' if ok > 1 else ''} promovido{'s' if ok > 1 else ''} al PAE",
            tipo="ok",
        )
        self._load_data()

    def resizeEvent(self, event):
        super().resizeEvent(event)
        if hasattr(self, '_confirm'):
            self._confirm._reposition()
        if hasattr(self, '_toast'):
            self._toast._reposition()

    def showEvent(self, event):
        super().showEvent(event)
        self._load_data()

    def refresh_period(self):
        """Llamado por main_window cuando el usuario cambia el período visualizado."""
        self._load_data()
